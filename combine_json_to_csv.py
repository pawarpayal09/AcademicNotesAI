import json
import csv
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ==========================================================
# PROJECT PATHS
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parent

STORAGE_DIR = PROJECT_ROOT / "storage"

CSV_OUTPUT_FILE = (
    STORAGE_DIR / "combined_student_data.csv"
)

EXCEL_OUTPUT_FILE = (
    STORAGE_DIR / "combined_student_data.xlsx"
)


# ==========================================================
# JSON FILES TO COMBINE
# ==========================================================

JSON_FILES = [
    "users.json",
    "chat_history.json",
    "favourites.json",
    "quiz_history.json",
    "image_study_history.json",
    "youtube_history.json",
    "activity_history.json",
]


# ==========================================================
# COMMON CSV / EXCEL COLUMNS
# ==========================================================

COMMON_COLUMNS = [
    "record_id",
    "source_file",
    "record_type",

    "user_uid",
    "user_name",
    "user_email",

    "date",
    "created_at",

    "topic",
    "question",
    "answer",

    "difficulty",
    "score",
    "total_questions",
    "percentage",

    "image_name",
    "instruction",

    "results_count",

    "chat_title",
    "message_count",

    "description",

    "sources",

    "email_verified",
    "disabled",

    "metadata_json",
]


# ==========================================================
# EXCEL SHEET MAPPING
# ==========================================================

SHEET_MAPPING = {
    "users": "Users",
    "chat_history": "Chats",
    "favourites": "Saved Notes",
    "quiz_history": "Quizzes",
    "image_study_history": "Image Study",
    "youtube_history": "YouTube",
    "activity_history": "Activity",
}


# ==========================================================
# EXCEL COLORS
# ==========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="2563EB"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="DBEAFE"
)

SUBHEADER_FONT = Font(
    color="075985",
    bold=True
)

THIN_BORDER = Border(
    bottom=Side(
        style="thin",
        color="D1D5DB"
    )
)


# ==========================================================
# SAFE JSON READ
# ==========================================================

def load_json_file(file_path):
    """
    Read one JSON file.

    Returns:
        list of records
    """

    if not file_path.exists():

        print(
            f"⚠️ File not found: {file_path.name}"
        )

        return []


    try:

        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(file)


        if isinstance(data, list):

            return data


        if isinstance(data, dict):

            return [data]


        print(
            f"⚠️ Unsupported JSON format: "
            f"{file_path.name}"
        )

        return []


    except json.JSONDecodeError as e:

        print(
            f"❌ Invalid JSON in "
            f"{file_path.name}: {e}"
        )

        return []


    except OSError as e:

        print(
            f"❌ Could not read "
            f"{file_path.name}: {e}"
        )

        return []


# ==========================================================
# VALUE CLEANER
# ==========================================================

def clean_value(value):
    """
    Convert dictionaries/lists into JSON strings
    so they can safely be stored inside CSV/Excel cells.
    """

    if value is None:

        return ""


    if isinstance(
        value,
        (dict, list)
    ):

        return json.dumps(
            value,
            ensure_ascii=False
        )


    return value


# ==========================================================
# GENERATE RECORD ID
# ==========================================================

def get_record_id(
    record,
    source_file,
    index
):

    possible_ids = [
        "id",
        "uid",
        "quiz_id",
        "activity_id",
        "search_id",
        "chat_id",
        "record_id",
    ]


    for key in possible_ids:

        value = record.get(
            key
        )

        if value not in (
            None,
            ""
        ):

            return str(
                value
            )


    return (
        f"{Path(source_file).stem}_"
        f"{index + 1}"
    )


# ==========================================================
# EXTRACT COMMON DATA
# ==========================================================

def convert_record(
    record,
    source_file,
    index
):
    """
    Convert one JSON record into one
    standardized tabular row.
    """

    row = {}


    # ------------------------------------------------------
    # BASIC INFORMATION
    # ------------------------------------------------------

    row["record_id"] = get_record_id(
        record,
        source_file,
        index
    )

    row["source_file"] = source_file

    row["record_type"] = Path(
        source_file
    ).stem


    # ------------------------------------------------------
    # USER INFORMATION
    # ------------------------------------------------------

    row["user_uid"] = record.get(
        "user_uid",
        record.get(
            "uid",
            ""
        )
    )

    row["user_name"] = record.get(
        "user_name",
        record.get(
            "name",
            ""
        )
    )

    row["user_email"] = record.get(
        "user_email",
        record.get(
            "email",
            ""
        )
    )


    # ------------------------------------------------------
    # TIME
    # ------------------------------------------------------

    row["date"] = clean_value(
        record.get(
            "date",
            ""
        )
    )

    row["created_at"] = clean_value(
        record.get(
            "created_at",
            ""
        )
    )


    # ------------------------------------------------------
    # ACADEMIC CONTENT
    # ------------------------------------------------------

    row["topic"] = clean_value(
        record.get(
            "topic",
            ""
        )
    )

    row["question"] = clean_value(
        record.get(
            "question",
            ""
        )
    )

    row["answer"] = clean_value(
        record.get(
            "answer",
            ""
        )
    )


    # ------------------------------------------------------
    # QUIZ
    # ------------------------------------------------------

    row["difficulty"] = clean_value(
        record.get(
            "difficulty",
            ""
        )
    )

    row["score"] = clean_value(
        record.get(
            "score",
            ""
        )
    )

    row["total_questions"] = clean_value(
        record.get(
            "total_questions",
            ""
        )
    )

    row["percentage"] = clean_value(
        record.get(
            "percentage",
            ""
        )
    )


    # ------------------------------------------------------
    # IMAGE STUDY
    # ------------------------------------------------------

    row["image_name"] = clean_value(
        record.get(
            "image_name",
            ""
        )
    )

    row["instruction"] = clean_value(
        record.get(
            "instruction",
            ""
        )
    )


    # ------------------------------------------------------
    # YOUTUBE
    # ------------------------------------------------------

    row["results_count"] = clean_value(
        record.get(
            "results_count",
            ""
        )
    )


    # ------------------------------------------------------
    # CHAT HISTORY
    # ------------------------------------------------------

    row["chat_title"] = clean_value(
        record.get(
            "title",
            record.get(
                "chat_title",
                ""
            )
        )
    )


    messages = record.get(
        "messages",
        []
    )


    if isinstance(
        messages,
        list
    ):

        row["message_count"] = len(
            messages
        )

    else:

        row["message_count"] = ""


    # ------------------------------------------------------
    # ACTIVITY
    # ------------------------------------------------------

    row["description"] = clean_value(
        record.get(
            "description",
            ""
        )
    )


    # ------------------------------------------------------
    # SOURCES
    # ------------------------------------------------------

    row["sources"] = clean_value(
        record.get(
            "sources",
            ""
        )
    )


    # ------------------------------------------------------
    # USER ACCOUNT FIELDS
    # ------------------------------------------------------

    row["email_verified"] = clean_value(
        record.get(
            "email_verified",
            ""
        )
    )

    row["disabled"] = clean_value(
        record.get(
            "disabled",
            ""
        )
    )


    # ------------------------------------------------------
    # UNKNOWN / ADDITIONAL DATA
    # ------------------------------------------------------

    known_keys = {
        "id",
        "uid",
        "quiz_id",
        "activity_id",
        "search_id",
        "chat_id",
        "record_id",

        "user_uid",
        "user_name",
        "user_email",

        "name",
        "email",

        "date",
        "created_at",

        "topic",
        "question",
        "answer",

        "difficulty",
        "score",
        "total_questions",
        "percentage",

        "image_name",
        "instruction",

        "results_count",

        "title",
        "chat_title",
        "messages",

        "description",
        "sources",

        "email_verified",
        "disabled",
    }


    extra_data = {
        key: value
        for key, value in record.items()
        if key not in known_keys
    }


    # ------------------------------------------------------
    # CHAT MESSAGES + EXTRA DATA
    # ------------------------------------------------------

    metadata = {}


    if extra_data:

        metadata.update(
            extra_data
        )


    if messages:

        metadata["messages"] = messages


    row["metadata_json"] = (
        json.dumps(
            metadata,
            ensure_ascii=False
        )
        if metadata
        else ""
    )


    return row


# ==========================================================
# CREATE DATAFRAME
# ==========================================================

def create_combined_dataframe():
    """
    Read all JSON files and create one
    standardized Pandas DataFrame.
    """

    STORAGE_DIR.mkdir(
        parents=True,
        exist_ok=True
    )


    all_rows = []


    print(
        "\n========================================"
    )

    print(
        " Academic Notes AI - JSON → DATASET"
    )

    print(
        "========================================\n"
    )


    for json_name in JSON_FILES:

        json_path = (
            STORAGE_DIR / json_name
        )


        print(
            f"📂 Reading: {json_name}"
        )


        records = load_json_file(
            json_path
        )


        print(
            f"   Records found: {len(records)}"
        )


        for index, record in enumerate(
            records
        ):

            if not isinstance(
                record,
                dict
            ):

                print(
                    f"⚠️ Skipping invalid "
                    f"record {index + 1} "
                    f"in {json_name}"
                )

                continue


            row = convert_record(
                record,
                json_name,
                index
            )


            all_rows.append(
                row
            )


    # ======================================================
    # CREATE DATAFRAME
    # ======================================================

    if all_rows:

        dataframe = pd.DataFrame(
            all_rows,
            columns=COMMON_COLUMNS
        )

    else:

        dataframe = pd.DataFrame(
            columns=COMMON_COLUMNS
        )


    # ======================================================
    # SORT DATA
    # ======================================================

    if not dataframe.empty:

        dataframe = (
            dataframe
            .sort_values(
                by=[
                    "user_uid",
                    "date",
                    "record_type"
                ],
                ascending=[
                    True,
                    False,
                    True
                ],
                na_position="last"
            )
            .reset_index(
                drop=True
            )
        )


    return dataframe


# ==========================================================
# WRITE CSV
# ==========================================================

def write_csv(
    dataframe
):
    """
    Write combined dataframe to CSV.
    """

    dataframe.to_csv(
        CSV_OUTPUT_FILE,
        index=False,
        encoding="utf-8-sig"
    )


# ==========================================================
# FORMAT EXCEL SHEET
# ==========================================================

def format_worksheet(
    worksheet,
    freeze_row=2
):
    """
    Apply professional formatting to an Excel sheet.
    """

    worksheet.freeze_panes = (
        f"A{freeze_row}"
    )


    # ------------------------------------------------------
    # Header formatting
    # ------------------------------------------------------

    for cell in worksheet[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = THIN_BORDER


    worksheet.row_dimensions[1].height = 24


    # ------------------------------------------------------
    # Cell alignment
    # ------------------------------------------------------

    for row in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


    # ------------------------------------------------------
    # Automatic column widths
    # ------------------------------------------------------

    for column_cells in worksheet.columns:

        column_letter = (
            get_column_letter(
                column_cells[0].column
            )
        )


        max_length = 0


        for cell in column_cells:

            try:

                value_length = len(
                    str(
                        cell.value
                        or ""
                    )
                )


                max_length = max(
                    max_length,
                    value_length
                )

            except Exception:

                pass


        # Keep widths reasonable.

        width = min(
            max(
                max_length + 2,
                12
            ),
            45
        )


        worksheet.column_dimensions[
            column_letter
        ].width = width


    # ------------------------------------------------------
    # Table formatting
    # ------------------------------------------------------

    if worksheet.max_row >= 2:

        table_ref = (
            f"A1:"
            f"{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )


        table = Table(
            displayName=(
                "Table_"
                + worksheet.title.replace(
                    " ",
                    "_"
                )
            ),
            ref=table_ref
        )


        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )


        table.tableStyleInfo = (
            table_style
        )


        worksheet.add_table(
            table
        )


# ==========================================================
# WRITE EXCEL WORKBOOK
# ==========================================================

def write_excel(
    combined_dataframe
):
    """
    Create a convenient multi-sheet Excel workbook.

    Sheets:
        Overview
        Users
        Chats
        Saved Notes
        Quizzes
        Image Study
        YouTube
        Activity
        Combined Dataset
    """

    with pd.ExcelWriter(
        EXCEL_OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        # ==================================================
        # OVERVIEW
        # ==================================================

        overview_data = [

            [
                "Academic Notes AI",
                "Combined Project Dataset"
            ],

            [
                "Generated On",
                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
            ],

            [
                "Total Records",
                len(
                    combined_dataframe
                )
            ],

            [
                "Total Columns",
                len(
                    COMMON_COLUMNS
                )
            ],

            [
                "Source JSON Files",
                len(
                    JSON_FILES
                )
            ],

        ]


        overview_df = pd.DataFrame(
            overview_data,
            columns=[
                "Metric",
                "Value"
            ]
        )


        overview_df.to_excel(
            writer,
            sheet_name="Overview",
            index=False
        )


        # ==================================================
        # INDIVIDUAL SHEETS
        # ==================================================

        for record_type, sheet_name in (
            SHEET_MAPPING.items()
        ):

            filtered = (
                combined_dataframe[
                    combined_dataframe[
                        "record_type"
                    ] == record_type
                ]
                .copy()
            )


            filtered.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )


        # ==================================================
        # COMBINED DATASET
        # ==================================================

        combined_dataframe.to_excel(
            writer,
            sheet_name="Combined Dataset",
            index=False
        )


    # ======================================================
    # POST-FORMATTING
    # ======================================================

    from openpyxl import load_workbook

    workbook = load_workbook(
        EXCEL_OUTPUT_FILE
    )


    # ------------------------------------------------------
    # Overview formatting
    # ------------------------------------------------------

    overview_sheet = workbook[
        "Overview"
    ]


    for cell in overview_sheet[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = Alignment(
            horizontal="center"
        )


    overview_sheet.column_dimensions[
        "A"
    ].width = 24


    overview_sheet.column_dimensions[
        "B"
    ].width = 35


    # ------------------------------------------------------
    # Format all other sheets
    # ------------------------------------------------------

    for worksheet in workbook.worksheets:

        if worksheet.title == "Overview":

            continue


        format_worksheet(
            worksheet
        )


    workbook.save(
        EXCEL_OUTPUT_FILE
    )


# ==========================================================
# MAIN FUNCTION
# ==========================================================

def combine_json_to_csv():
    """
    Main function.

    Every run:
        1. Reads latest JSON files.
        2. Rebuilds combined CSV.
        3. Rebuilds Excel workbook.

    Therefore newly added records are always included.
    """

    dataframe = (
        create_combined_dataframe()
    )


    # ======================================================
    # CSV
    # ======================================================

    write_csv(
        dataframe
    )


    # ======================================================
    # EXCEL
    # ======================================================

    write_excel(
        dataframe
    )


    # ======================================================
    # FINAL SUMMARY
    # ======================================================

    print(
        "\n========================================"
    )

    print(
        "✅ DATASET CREATED SUCCESSFULLY"
    )

    print(
        "========================================"
    )


    print(
        f"📄 CSV:"
        f"\n   {CSV_OUTPUT_FILE}"
    )


    print(
        f"\n📊 Excel:"
        f"\n   {EXCEL_OUTPUT_FILE}"
    )


    print(
        f"\n📈 Total records: "
        f"{len(dataframe)}"
    )


    print(
        f"📋 Total columns: "
        f"{len(COMMON_COLUMNS)}"
    )


    print(
        "\n========================================"
    )


    print(
        "✅ JSON → CSV + Excel completed"
    )


    print(
        "========================================\n"
    )


    return (
        CSV_OUTPUT_FILE,
        EXCEL_OUTPUT_FILE
    )


# ==========================================================
# RUN SCRIPT
# ==========================================================

if __name__ == "__main__":

    combine_json_to_csv()