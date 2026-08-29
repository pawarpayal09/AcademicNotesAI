import os
from io import BytesIO

import streamlit as st

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from firebase_manager import (
    require_login,
    is_authenticated,
    get_current_user,
    logout_user
)

from data_science.dataset_manager import (
    get_all_data
)


# ==========================================================
# PAGE CONFIG
# ==========================================================

st.set_page_config(
    page_title="Dataset",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ==========================================================
# LOGIN REQUIRED
# ==========================================================

require_login()


# ==========================================================
# LOAD SHARED CSS
# IMPORTANT:
# CSS MUST BE INSIDE <style> TAGS
# ==========================================================

def load_css():

    css_path = os.path.join(
        os.path.dirname(
            os.path.dirname(__file__)
        ),
        "css",
        "style.css"
    )

    if not os.path.exists(css_path):
        return

    try:

        with open(
            css_path,
            "r",
            encoding="utf-8"
        ) as file:

            css_content = file.read()

        st.markdown(
            f"""
            <style>
            {css_content}
            </style>
            """,
            unsafe_allow_html=True
        )

    except OSError:
        pass


load_css()


# ==========================================================
# CURRENT USER
# ==========================================================

if is_authenticated():

    current_user = get_current_user()

else:

    st.stop()


if not current_user:

    st.stop()


# ==========================================================
# TOP-RIGHT USER PROFILE
# ==========================================================

profile_spacer, profile_area = st.columns(
    [5.8, 1.8],
    gap="small"
)


with profile_area:

    with st.container(
        key="home_user_profile"
    ):

        # -----------------------------------------
        # USER NAME
        # -----------------------------------------

        st.markdown(
            f"""
            <div class="home-profile-name">
                👤 {current_user['name']}
            </div>
            """,
            unsafe_allow_html=True
        )

        # -----------------------------------------
        # USER EMAIL
        # -----------------------------------------

        st.markdown(
            f"""
            <div class="home-profile-email">
                {current_user['email']}
            </div>
            """,
            unsafe_allow_html=True
        )

        # -----------------------------------------
        # LOGOUT
        # -----------------------------------------

        if st.button(
            "🚪 Logout",
            key="home_logout_button",
            use_container_width=True
        ):

            logout_user()

            st.rerun()



# =====================================================
# SIDEBAR
# =====================================================

with st.sidebar:

    st.markdown(
        """
        # 🧑‍🎓 StudyNova

        ### Learn Smarter with AI
        """
    )

    st.divider()

    st.info(
        """
### 🤖 About

An AI-powered academic assistant that helps students:

- 📚 Search academic notes
- 📄 Chat with uploaded PDFs
- 🖼️ Understand study images
- 🧠 Generate practice quizzes
- 🎥 Find educational YouTube resources
- 📌 Save important AI answers
- 📊 Track learning activity
- 📈 View personal study progress
"""
    )

    st.divider()

    st.markdown(
        "### 🚀 Technologies"
    )

    st.success(
        "🤖 Google Gemini"
    )

    st.success(
        "🦜 LangChain"
    )

    st.success(
        "📚 FAISS"
    )

    st.success(
        "🧠 RAG"
    )

    st.success(
        "🔥 Firebase"
    )

    st.success(
        "🎥 YouTube API"
    )

    st.success(
        "⚡ Streamlit"
    )

    st.divider()

    st.markdown(
        "### 👩‍💻 Developer"
    )

    st.info(
        """
**Payal Pawar**

🎓 MCA Student

StudyNova

Academic Notes AI

Version 1.0
"""
    )

# ==========================================================
# PAGE HEADER
# ==========================================================

st.title(
    "📊 Academic Notes AI — Dataset & Analysis"
)

st.write(
    "View the APIs used in the project and "
    "the complete user activity dataset."
)

st.caption(
    f"Currently logged in as: "
    f"{current_user.get('email', '')}"
)

st.divider()


# ==========================================================
# LOAD COMPLETE DATASET
# IMPORTANT:
# THIS LOADS ALL USERS' RECORDS
# ==========================================================

with st.spinner(
    "📥 Loading complete activity dataset..."
):

    dataframe = get_all_data()


# ==========================================================
# API INFORMATION
# ==========================================================

st.markdown(
    "## 🔌 APIs Used in Academic Notes AI"
)

st.write(
    "The following APIs are used in different "
    "modules of the Academic Notes AI project."
)


# ==========================================================
# API CARDS
# ==========================================================

api1, api2, api3 = st.columns(
    3,
    gap="small"
)


# ==========================================================
# MAIN CHAT API 1
# ==========================================================

with api1:

    with st.container(
        border=True
    ):

        st.markdown(
            "### Main Chatbot API"
        )

        st.caption(
            "GOOGLE_GEMINI_API_KEY"
        )

        st.markdown(
            "**Used in:**"
        )

        st.write(
            "Main Academic Chatbot"
        )

        st.markdown(
            "**Purpose:**"
        )

        st.write(
            "Generates AI answers using the "
            "student's question and relevant "
            "academic context retrieved from RAG."
        )


# ==========================================================
# IMAGE STUDY API
# ==========================================================

with api2:

    with st.container(
        border=True
    ):

        st.markdown(
            "### Image Study API"
        )

        st.caption(
            "IMAGE_STUDY_API_KEY"
        )

        st.markdown(
            "**Used in:**"
        )

        st.write(
            "Image Study Assistant"
        )

        st.markdown(
            "**Purpose:**"
        )

        st.write(
            "Analyzes uploaded study images, "
            "diagrams and screenshots and "
            "explains them in simple language."
        )


# ==========================================================
# YOUTUBE DATA API
# ==========================================================

with api3:

    with st.container(
        border=True
    ):

        st.markdown(
            "### YouTube Data API"
        )

        st.caption(
            "YOUTUBE_DATA_API_KEY"
        )

        st.markdown(
            "**Used in:**"
        )

        st.write(
            "YouTube Learning Resources"
        )

        st.markdown(
            "**Purpose:**"
        )

        st.write(
            "Searches real educational videos "
            "related to the selected academic topic "
            "and helps students learn."
        )


# ==========================================================
# ROW 2
# ==========================================================

st.markdown("")


api4, api5, empty_column = st.columns(
    3,
    gap="small"
)


# ==========================================================
# FIREBASE API
# ==========================================================

with api4:

    with st.container(
        border=True
    ):

        st.markdown(
            "### Firebase API"
        )

        st.caption(
            "FIREBASE_API_KEY"
        )

        st.markdown(
            "**Used in:**"
        )

        st.write(
            "Authentication and user management"
        )

        st.markdown(
            "**Purpose:**"
        )

        st.write(
            "Handles signup, login, password reset "
            "and user identity through Firebase."
        )


# ==========================================================
# MAIN CHAT API 2
# ==========================================================

with api5:

    with st.container(
        border=True
    ):

        st.markdown(
            "### Main Chatbot API 2"
        )

        st.caption(
            "GOOGLE_GEMINI_API_KEY"
        )

        st.markdown(
            "**Used in:**"
        )

        st.write(
            "Main Academic Chatbot"
        )

        st.markdown(
            "**Purpose:**"
        )

        st.write(
            "Provides the second configured "
            "Gemini credential for the main "
            "chat generation flow."
        )


# ==========================================================
# DATASET SECTION
# ==========================================================

st.divider()

st.markdown(
    "## 📋 User Activity Dataset"
)

st.caption(
    "Complete activity records of all users "
    "from the beginning of the stored dataset "
    "up to the latest available record."
)


# ==========================================================
# NULL / BLANK VALUE HELPER
# ==========================================================

def prepare_display_dataframe(
    source_dataframe
):
    """
    Create a clean display copy.

    Blank, empty and missing values are
    displayed as 'NULL'.

    The original DataFrame is not modified.
    """

    clean_dataframe = (
        source_dataframe.copy()
    )


    for column in clean_dataframe.columns:

        clean_dataframe[column] = (
            clean_dataframe[column]
            .apply(
                lambda value:
                "NULL"
                if (
                    value is None
                    or (
                        isinstance(
                            value,
                            str
                        )
                        and not value.strip()
                    )
                    or (
                        not isinstance(
                            value,
                            (
                                list,
                                dict,
                                tuple
                            )
                        )
                        and str(value).lower()
                        in {
                            "nan",
                            "nat",
                            "none"
                        }
                    )
                )
                else value
            )
        )


    return clean_dataframe


# ==========================================================
# DATASET DISPLAY
# ==========================================================

if dataframe.empty:

    st.info(
        """
        📭 No activity records are available yet.

        Start using Chatbot, Quiz, Image Study,
        YouTube or Saved Notes to generate data.
        """
    )

else:

    # ------------------------------------------------------
    # Preferred column order
    # ------------------------------------------------------

    preferred_columns = [

        "record_id",

        "record_type",

        "source_file",

        "user_uid",

        "user_name",

        "user_email",

        "date",

        "created_at",

        "topic",

        "question",

        "answer",

        "difficulty",

        "score",

        "total_questions",

        "percentage",

        "image_name",

        "instruction",

        "results_count",

        "chat_title",

        "message_count",

        "description",

        "sources",

        "email_verified",

        "disabled",

        "question_length",

        "answer_length",

        "has_question",

        "has_answer",

        "has_source",

        "activity_hour"
    ]


    # ------------------------------------------------------
    # Existing columns
    # ------------------------------------------------------

    available_columns = [

        column
        for column in preferred_columns
        if column in dataframe.columns
    ]


    # ------------------------------------------------------
    # Any additional backend columns
    # ------------------------------------------------------

    remaining_columns = [

        column
        for column in dataframe.columns
        if column not in available_columns
    ]


    final_columns = (
        available_columns
        +
        remaining_columns
    )


    display_dataframe = (
        dataframe[
            final_columns
        ]
        .copy()
    )


    # ------------------------------------------------------
    # Replace blank/null values
    # ------------------------------------------------------

    display_dataframe = (
        prepare_display_dataframe(
            display_dataframe
        )
    )


    # ======================================================
    # DISPLAY COMPLETE DATASET
    # ======================================================

    st.dataframe(
        display_dataframe,
        use_container_width=True,
        hide_index=True,
        height=600
    )


    # ======================================================
    # DATASET COUNT
    # ======================================================

    st.caption(
        f"Total records: "
        f"{len(display_dataframe)}"
        f" • Total columns: "
        f"{len(display_dataframe.columns)}"
    )


    # ======================================================
    # CREATE EXCEL FILE
    # ======================================================

    def create_excel_file(
        dataset
    ):
        """
        Create a formatted Excel file in memory.
        """

        import pandas as pd

        output = BytesIO()


        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:

            dataset.to_excel(
                writer,
                sheet_name="Fused Dataset",
                index=False
            )


            worksheet = (
                writer.sheets[
                    "Fused Dataset"
                ]
            )


            # --------------------------------------------------
            # Header
            # --------------------------------------------------

            header_fill = PatternFill(
                fill_type="solid",
                fgColor="2563EB"
            )


            header_font = Font(
                color="FFFFFF",
                bold=True
            )


            thin_border = Border(
                bottom=Side(
                    style="thin",
                    color="D1D5DB"
                )
            )


            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cell.border = thin_border


            # --------------------------------------------------
            # Freeze header
            # --------------------------------------------------

            worksheet.freeze_panes = "A2"


            # --------------------------------------------------
            # Filter
            # --------------------------------------------------

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )


            # --------------------------------------------------
            # Header row height
            # --------------------------------------------------

            worksheet.row_dimensions[
                1
            ].height = 28


            # --------------------------------------------------
            # Cell alignment
            # --------------------------------------------------

            for row in worksheet.iter_rows(
                min_row=2
            ):

                for cell in row:

                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )


            # --------------------------------------------------
            # Column widths
            # --------------------------------------------------

            for column_cells in worksheet.columns:

                column_letter = (
                    get_column_letter(
                        column_cells[0].column
                    )
                )


                max_length = 0


                for cell in column_cells:

                    try:

                        value_length = len(
                            str(
                                cell.value
                                if cell.value is not None
                                else ""
                            )
                        )


                        max_length = max(
                            max_length,
                            value_length
                        )

                    except Exception:

                        pass


                width = min(
                    max(
                        max_length + 2,
                        12
                    ),
                    45
                )


                worksheet.column_dimensions[
                    column_letter
                ].width = width


            # --------------------------------------------------
            # Excel Table
            # --------------------------------------------------

            if worksheet.max_row >= 2:

                from openpyxl.worksheet.table import (
                    Table,
                    TableStyleInfo
                )


                table_ref = (
                    f"A1:"
                    f"{get_column_letter(worksheet.max_column)}"
                    f"{worksheet.max_row}"
                )


                table = Table(
                    displayName="FusedDatasetTable",
                    ref=table_ref
                )


                table_style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )


                table.tableStyleInfo = (
                    table_style
                )


                worksheet.add_table(
                    table
                )


        output.seek(0)

        return output.getvalue()


    # ======================================================
    # DOWNLOAD FUSED DATASET
    # ======================================================

    excel_file = create_excel_file(
        display_dataframe
    )


    st.download_button(
        label="📥 Download Fussed Dataset",
        data=excel_file,
        file_name="AcademicNotesAI_Fused_Dataset.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=False,
        type="primary",
        key="download_fused_dataset"
    )


# ==========================================================
# FOOTER
# ==========================================================

st.divider()


footer1, footer2, footer3 = st.columns(
    3
)


with footer1:

    st.caption(
        "📊 Academic Notes AI Dataset"
    )


with footer2:

    st.caption(
        "⚡ Python • Pandas • Streamlit"
    )


with footer3:

    st.caption(
        "👩‍💻 Developed by Payal Pawar"
    )